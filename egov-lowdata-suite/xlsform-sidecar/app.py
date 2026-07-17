import io
from typing import Dict, List, Any

from flask import Flask, request, send_file, jsonify
from openpyxl import Workbook

app = Flask(__name__)


@app.route("/", methods=["GET"])
def index():
    """Friendly landing page for manual health checks."""
    return jsonify(
        {
            "service": "xlsform-sidecar",
            "description": "POST a JSON form spec with `fields` to /generate to receive an XLSForm.",
            "endpoints": {"health": "/health", "generate": "/generate"},
        }
    )


@app.route("/health", methods=["GET"])
def health():
    """Simple healthcheck endpoint."""
    return jsonify(status="ok")


def build_workbook(spec: Dict[str, Any]) -> io.BytesIO:
    title = spec.get("title") or spec.get("name") or "Generated Form"
    form_id = spec.get("form_id") or spec.get("name") or title.lower().replace(" ", "_")
    fields: List[Dict[str, Any]] = spec.get("fields", [])
    if not isinstance(fields, list) or not fields:
        raise ValueError("fields 배열이 비어있습니다.")

    wb = Workbook()
    survey_ws = wb.active
    survey_ws.title = "survey"
    survey_ws.append(["type", "name", "label::English", "required"])

    choices_ws = wb.create_sheet("choices")
    choices_ws.append(["list_name", "name", "label::English"])
    has_choices = False

    for field in fields:
        if not isinstance(field, dict):
            raise ValueError("각 field는 객체 형태여야 합니다.")
        field_type = field.get("type")
        name = field.get("name")
        label = field.get("label", name or "")

        if not field_type or not name:
            raise ValueError("각 field에는 type과 name이 필요합니다.")

        required = "yes" if field.get("required") else ""

        if field_type in ("select_one", "select_multiple"):
            choices = field.get("choices", [])
            if not isinstance(choices, list) or not choices:
                raise ValueError(f"{name} 필드의 choices가 비어있습니다.")
            list_name = field.get("list_name", f"{name}_list")
            has_choices = True
            for choice in choices:
                if not isinstance(choice, dict):
                    raise ValueError(f"{name} 필드의 choice가 객체 형태가 아닙니다.")
                choice_name = choice.get("name")
                choice_label = choice.get("label", choice_name or "")
                if not choice_name:
                    raise ValueError(f"{name} 필드의 choice에 name이 필요합니다.")
                choices_ws.append([list_name, choice_name, choice_label])
            xls_type = f"{field_type} {list_name}"
        else:
            xls_type = field_type

        survey_ws.append([xls_type, name, label, required])

    if not has_choices:
        # keep header row but nothing else
        pass

    settings_ws = wb.create_sheet("settings")
    settings_ws.append(["form_title", "form_id"])
    settings_ws.append([title, form_id])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route("/generate", methods=["POST"])
def generate_xlsform():
    """Converts a simplified JSON form specification into an XLSForm (XLSX bytes)."""
    try:
        spec = request.get_json(force=True)
    except Exception:
        return jsonify({"error": "Invalid JSON"}), 400

    if not isinstance(spec, dict):
        return jsonify({"error": "JSON body must be an object"}), 400

    try:
        output = build_workbook(spec)
        filename = f"{spec.get('form_id', spec.get('name', 'form')).replace(' ', '_')}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except ValueError as ve:
        return jsonify({"error": str(ve)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5001)
